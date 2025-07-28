
#------------------------------------------------------------#
import torch
import streamlit as st
from PIL import Image
import os
import cv2
import numpy as np
import fitz  # PyMuPDF for PDF handling
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from combined import process_invoice, InvoiceProcessor
from image_path import Output_Text_File, Image_VietOCR_Folder
from Praser import process_invoice_file
import time

def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

local_css("style.css")

# Initialize session state variables
def init_session_state():
    if 'ocr_results' not in st.session_state:
        st.session_state.ocr_results = {}
    if 'df_results' not in st.session_state:
        st.session_state.df_results = {}
    if 'full_texts' not in st.session_state:
        st.session_state.full_texts = {}
    if 'processed_images' not in st.session_state:
        st.session_state.processed_images = {}
    if 'parsed_data' not in st.session_state:
        st.session_state.parsed_data = {}
    if 'selected_file_idx' not in st.session_state:
        st.session_state.selected_file_idx = 0
    if 'recognition_clicked' not in st.session_state:
        st.session_state.recognition_clicked = False
    if 'current_page' not in st.session_state:
        st.session_state.current_page = 0
    if 'current_pdf' not in st.session_state:
        st.session_state.current_pdf = None
    if 'total_pages' not in st.session_state:
        st.session_state.total_pages = 0
    if 'batch_processing' not in st.session_state:
        st.session_state.batch_processing = False
    if 'batch_current_idx' not in st.session_state:
        st.session_state.batch_current_idx = 0
    if 'batch_progress' not in st.session_state:
        st.session_state.batch_progress = 0
    if 'batch_status' not in st.session_state:
        st.session_state.batch_status = ""

init_session_state()

def process_single_file(file, file_idx):
    """Process a single file for OCR recognition"""
    try:
        filename = file.name
        file_key = f"{filename}_{file_idx}"
        
        # Skip if already processed
        if file_key in st.session_state.ocr_results:
            return True, f"File {filename} already processed"
        
        # Read and prepare the image
        file.seek(0)  # Reset file pointer
        is_pdf = filename.lower().endswith('.pdf')
        
        if is_pdf:
            # Handle PDF files
            pdf_content = file.read()
            pdf_stream = BytesIO(pdf_content)
            pdf_doc = fitz.open(stream=pdf_stream, filetype="pdf")
            
            # Process first page of PDF for now
            page = pdf_doc.load_page(0)
            pix = page.get_pixmap()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_np = np.array(img)
            pdf_doc.close()
        else:
            # Handle regular images
            image = Image.open(file)
            img_np = np.array(image)
        
        # Save temporary image for processing
        file_basename = os.path.splitext(os.path.basename(filename))[0]
        temp_image_path = os.path.join(Image_VietOCR_Folder, f"{file_basename}_temp.jpg")
        
        # Ensure directory exists
        os.makedirs(Image_VietOCR_Folder, exist_ok=True)
        
        # Convert and save image
        if len(img_np.shape) == 3 and img_np.shape[2] == 3:  # RGB
            cv2.imwrite(temp_image_path, cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR))
        else:  # Grayscale or other
            cv2.imwrite(temp_image_path, img_np)
        
        # Process with OCR
        processor = InvoiceProcessor()
        result = processor.process_image(Image.fromarray(img_np))
        
        # Save OCR result to text file
        temp_text_path = os.path.join(Image_VietOCR_Folder, f"{file_basename}.txt")
        with open(temp_text_path, 'w', encoding='utf-8') as f:
            if isinstance(result, dict) and 'full_text' in result:
                f.write(result['full_text'])
            else:
                f.write(str(result))
        
        # Process with Parser
        output_excel = os.path.join(Image_VietOCR_Folder, f"{file_basename}_parsed.xlsx")
        process_invoice_file(temp_text_path, output_excel)
        
        # Read parsed data
        parsed_data = {}
        if os.path.exists(output_excel):
            try:
                df_parsed = pd.read_excel(output_excel).astype(str)
                parsed_data = df_parsed.to_dict('records')[0] if not df_parsed.empty else {}
            except Exception as e:
                parsed_data = {"Error": f"Could not read Excel file: {str(e)}"}
        else:
            parsed_data = {"Info": "Excel file not generated"}
        
        # Store results in session state
        if isinstance(result, dict):
            st.session_state.ocr_results[file_key] = result.get('text_blocks', [])
            st.session_state.full_texts[file_key] = result.get('full_text', '')
        else:
            st.session_state.ocr_results[file_key] = [str(result)]
            st.session_state.full_texts[file_key] = str(result)
        
        st.session_state.parsed_data[file_key] = parsed_data
        
        # Store processed image info
        st.session_state.processed_images[file_key] = {
            'image': img_np,
            'filename': filename,
            'is_pdf': is_pdf
        }
        
        # Clean up temporary files
        for temp_file in [temp_image_path, temp_text_path]:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass  # Ignore cleanup errors
        
        return True, f"Successfully processed {filename}"
        
    except Exception as e:
        return False, f"Error processing {filename}: {str(e)}"

# ===== Upload file =====
st.sidebar.title("🚀 Bắt Đầu Nhận Diện")

# Allow multiple file uploads
uploaded_files = st.sidebar.file_uploader(
    "Chọn tệp ảnh hoặc PDF",
    type=["jpg", "jpeg", "png", "pdf"],
    accept_multiple_files=True,
    key="file_uploader"
)

# ===== File list in sidebar =====
if uploaded_files:
    st.sidebar.subheader("Danh sách tệp đã tải lên")
    selected_file_idx = st.session_state.get('selected_file_idx', 0)
    
    # Display clickable file list
    for i, file in enumerate(uploaded_files):
        # Create a container for each file item
        file_container = st.sidebar.container()
        
        # Use columns to create a more interactive layout
        col_icon, col_name = file_container.columns([0.1, 0.9])
        
        with col_icon:
            # Show different icons based on processing status
            file_key = f"{file.name}_{i}"
            if file_key in st.session_state.ocr_results:
                st.markdown("✅")  # Processed
            elif st.session_state.batch_processing and i == st.session_state.batch_current_idx:
                st.markdown("🔄")  # Currently processing
            else:
                st.markdown("📄")  # Not processed
            
        with col_name:
            # Create a button that looks like text
            if st.button(
                file.name,
                key=f"file_btn_{i}",
                use_container_width=True,
                type="primary" if i == selected_file_idx else "secondary",
                help=f"Chọn để xem và nhận dạng"
            ):
                # Update selected file index
                st.session_state.selected_file_idx = i
                # Reset recognition state when changing files
                st.session_state.recognition_clicked = False
                st.session_state.current_page = 0  # Reset to first page for PDFs
                st.rerun()
                
        # Add a small separator between files
        if i < len(uploaded_files) - 1:
            st.sidebar.markdown("<hr style='margin: 0.5rem 0; border-color: #555;'>", unsafe_allow_html=True)

    st.sidebar.markdown("---")
    
    # Batch processing button and status
    col1, col2 = st.sidebar.columns([1, 1])
    
    with col1:
        if not st.session_state.batch_processing:
            if st.button("🔍 Nhận dạng tất cả", use_container_width=True):
                st.session_state.batch_processing = True
                st.session_state.batch_current_idx = 0
                st.session_state.batch_progress = 0
                st.session_state.batch_status = "Bắt đầu xử lý..."
                st.rerun()
        else:
            if st.button("⏹️ Dừng xử lý", use_container_width=True):
                st.session_state.batch_processing = False
                st.session_state.batch_status = "Đã dừng xử lý"
                st.rerun()
    
    with col2:
        # Show processing progress
        if st.session_state.batch_processing:
            progress = st.session_state.batch_current_idx / len(uploaded_files)
            st.progress(progress)
        else:
            # Count processed files
            processed_count = sum(1 for i, file in enumerate(uploaded_files) 
                                if f"{file.name}_{i}" in st.session_state.ocr_results)
            st.write(f"✅ {processed_count}/{len(uploaded_files)}")
    
    # Show batch processing status
    if st.session_state.batch_status:
        st.sidebar.info(st.session_state.batch_status)

# ===== Batch Processing Logic =====
if st.session_state.batch_processing and uploaded_files:
    current_idx = st.session_state.batch_current_idx
    
    if current_idx < len(uploaded_files):
        file = uploaded_files[current_idx]
        filename = file.name
        
        # Update status
        st.session_state.batch_status = f"Đang xử lý ({current_idx + 1}/{len(uploaded_files)}): {filename}"
        
        # Create a progress container in main area
        progress_container = st.container()
        with progress_container:
            st.markdown("### 🔄 Đang thực hiện nhận dạng hàng loạt...")
            
            # Overall progress
            overall_progress = st.progress(current_idx / len(uploaded_files))
            st.write(f"Tiến độ: {current_idx}/{len(uploaded_files)} tệp đã hoàn thành")
            
            # Current file status
            st.info(f"Đang xử lý: **{filename}**")
            
            # Processing animation
            with st.spinner(f"Nhận dạng văn bản từ {filename}..."):
                # Process the current file
                success, message = process_single_file(file, current_idx)
                
                if success:
                    st.success(f"✅ {message}")
                    time.sleep(0.5)  # Brief pause to show success
                else:
                    st.error(f"❌ {message}")
                    time.sleep(1)  # Longer pause to show error
        
        # Move to next file
        st.session_state.batch_current_idx = current_idx + 1
        st.rerun()
        
    else:
        # Batch processing completed
        st.session_state.batch_processing = False
        st.session_state.batch_status = f"✅ Hoàn thành! Đã xử lý {len(uploaded_files)} tệp."
        
        # Show completion message
        st.success("🎉 Đã hoàn thành nhận dạng tất cả các tệp!")
        
        # Create summary of results
        with st.expander("📊 Tóm tắt kết quả xử lý", expanded=True):
            summary_data = []
            for i, file in enumerate(uploaded_files):
                file_key = f"{file.name}_{i}"
                status = "Thành công" if file_key in st.session_state.ocr_results else "Thất bại"
                summary_data.append({
                    "STT": i + 1,
                    "Tên tệp": file.name,
                    "Trạng thái": status
                })
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True)
        
        time.sleep(2)  # Show completion message briefly
        st.rerun()

# ===== Main Content Area =====
col1, col2 = st.columns([9, 1])

with col1:
    st.markdown(
        """
        <div style='border: 2px solid #0B5394; padding: 20px; border-radius: 10px; background-color: #f0f8ff;'>
            <h1 style='text-align: center; font-size: 40px; color: #0B5394; margin: 0;'>
                NHẬN DIỆN HÓA ĐƠN TỰ ĐỘNG
            </h1>
        </div>
        """,
        unsafe_allow_html=True
    )

with col2:
    st.image("C:/Users/Admin/Desktop/OCR/Logo-HCMIU.png", use_column_width="auto")
# Don't show individual file processing during batch processing
if not st.session_state.batch_processing and uploaded_files:
    # Use the selected file index from session state, but ensure it's within bounds
    idx = st.session_state.get('selected_file_idx', 0)
    if idx >= len(uploaded_files) or idx < 0:
        idx = 0  # Reset to first file if index is out of range
        st.session_state.selected_file_idx = idx
    
    file = uploaded_files[idx]
    filename = file.name
    file_key = f"{filename}_{idx}"
    is_pdf = filename.lower().endswith('.pdf')
    
    # Store uploaded file data in session state
    if file_key not in st.session_state.processed_images or st.session_state.current_pdf != file_key:
        if is_pdf:
            try:
                # For PDFs, store the file content and metadata
                pdf_content = file.read()
                pdf_stream = BytesIO(pdf_content)
                pdf_doc = fitz.open(stream=pdf_stream, filetype="pdf")
                total_pages = len(pdf_doc)
                
                # Store PDF content and metadata
                st.session_state.current_pdf = file_key
                st.session_state.total_pages = total_pages
                st.session_state.current_page = min(st.session_state.get('current_page', 0), max(0, total_pages - 1))
                
                # Load the current page
                page = pdf_doc.load_page(st.session_state.current_page)
                pix = page.get_pixmap()
                
                # Convert to PIL Image
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                img_np = np.array(img)
                
                # Store the processed image
                st.session_state.processed_images[file_key] = {
                    'image': img_np,
                    'filename': filename,
                    'is_pdf': True,
                    'pdf_content': pdf_content,
                    'current_page': st.session_state.current_page,
                    'total_pages': total_pages
                }
                
                pdf_doc.close()
                
            except Exception as e:
                st.error(f"Lỗi khi xử lý file PDF: {str(e)}")
                st.session_state.processed_images[file_key] = {
                    'image': None,
                    'filename': filename,
                    'is_pdf': True,
                    'error': str(e)
                }
        else:
            # For regular images
            try:
                image = Image.open(file)
                img_np = np.array(image)
                
                st.session_state.processed_images[file_key] = {
                    'image': img_np,
                    'filename': filename,
                    'is_pdf': False
                }
            except Exception as e:
                st.error(f"Lỗi khi xử lý ảnh: {str(e)}")
                st.session_state.processed_images[file_key] = {
                    'image': None,
                    'filename': filename,
                    'is_pdf': False,
                    'error': str(e)
                }
    st.markdown("---")
    # Create main content columns
    col1, col2 = st.columns([1, 2])
    
    # Display image and controls
    with col2:
        file_data = st.session_state.processed_images[file_key]
        is_processed = file_key in st.session_state.ocr_results
        
        # PDF navigation controls
        if not is_processed and file_data.get('is_pdf', False):
            nav_container = st.container()
            with nav_container:
                col_prev, col_next, col_page = st.columns([1.2, 1.2, 1])
                
                with col_prev:
                    if st.button("⬅️ Trang Trước", 
                               key=f"prev_{idx}", 
                               disabled=st.session_state.current_page <= 0):
                        st.session_state.current_page = max(0, st.session_state.current_page - 1)
                        st.rerun()
                
                with col_next:
                    if st.button("Trang Sau ➡️", 
                               key=f"next_{idx}", 
                               disabled=st.session_state.current_page >= file_data['total_pages'] - 1):
                        st.session_state.current_page = min(file_data['total_pages'] - 1, st.session_state.current_page + 1)
                        st.rerun()
                
                with col_page:
                    page_num = st.number_input("Trang", 
                                            min_value=1, 
                                            max_value=file_data['total_pages'], 
                                            value=st.session_state.current_page + 1,
                                            key=f"page_input_{idx}",
                                            label_visibility="collapsed")
                
                if page_num - 1 != st.session_state.current_page:
                    st.session_state.current_page = page_num - 1
                    st.rerun()
            
            st.caption(f"Trang {st.session_state.current_page + 1}/{file_data['total_pages']}")
            
            # Reload PDF page if needed
            if 'pdf_content' in file_data:
                pdf_doc = fitz.open(stream=file_data['pdf_content'], filetype="pdf")
                page = pdf_doc.load_page(st.session_state.current_page)
                pix = page.get_pixmap()
                img_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
                file_data['image'] = img_np
                pdf_doc.close()
        
        # Display the image
        if file_data['image'] is not None:
            st.image(file_data['image'], 
                    caption=f"{filename} (Trang {st.session_state.current_page + 1}/{file_data.get('total_pages', 1)})" 
                            if file_data.get('is_pdf', False) else f"{filename}",
                    use_column_width=True)
        
        # Individual recognition button
        if st.button(f"🔍 Nhận dạng văn bản" if not is_processed else "✅ Đã nhận dạng", 
                   key=f"btn_{idx}",
                   disabled=is_processed,
                   help="Nhấn để bắt đầu nhận dạng văn bản" if not is_processed else "Văn bản đã được nhận dạng"):
            
            with st.spinner("Đang xử lý..."):
                    try:
                        # First, save the image to a temporary file
                        file_basename = os.path.splitext(os.path.basename(filename))[0]
                        temp_image_path = os.path.join(Image_VietOCR_Folder, f"{file_basename}.jpg")
                        cv2.imwrite(temp_image_path, cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR))
                        
                        # Process the image using the existing InvoiceProcessor
                        processor = InvoiceProcessor()
                        result = processor.process_image(Image.fromarray(img_np))
                        
                        # Save the OCR text to a temporary file for the Parser
                        temp_text_path = os.path.join(Image_VietOCR_Folder, f"{file_basename}.txt")
                        with open(temp_text_path, 'w', encoding='utf-8') as f:
                            f.write(result['full_text'])
                        
                        # Process the text file with the Parser
                        output_excel = os.path.join(Image_VietOCR_Folder, f"{file_basename}_parsed.xlsx")
                        process_invoice_file(temp_text_path, output_excel)
                        
                        # Read the parsed data from Excel
                        if os.path.exists(output_excel):
                            df_parsed = pd.read_excel(output_excel)
                            # Convert all columns to string for display
                            df_parsed = df_parsed.astype(str)
                        else:
                            df_parsed = pd.DataFrame({"Thông tin": ["Không thể đọc dữ liệu từ file Excel"]})
                        
                        # Store results in session state
                        st.session_state.ocr_results[file_key] = result['text_blocks']
                        st.session_state.df_results[file_key] = df_parsed
                        st.session_state.full_texts[file_key] = result['full_text']
                        
                        # Store the parsed data in session state for display
                        st.session_state.parsed_data[file_key] = df_parsed.to_dict('records')[0] if not df_parsed.empty else {}
                        
                        st.experimental_rerun()
                        
                    except Exception as e:
                        st.error(f"Lỗi khi xử lý ảnh: {str(e)}")
                        st.session_state.recognition_clicked = False

    # ===== Show recognition results if available =====
    if file_key in st.session_state.ocr_results:
        # Show the original image from session state
        if file_key in st.session_state.processed_images and 'image' in st.session_state.processed_images[file_key]:
            st.image(st.session_state.processed_images[file_key]['image'], 
                   caption=f"Kết quả nhận dạng: {filename}", 
                   use_column_width=True)
        
        # Get OCR results
        text_blocks = st.session_state.ocr_results[file_key]
        full_text = st.session_state.full_texts[file_key]
        
        # 1. Show individual OCR blocks first
        if text_blocks:
            with st.expander("🔍 Các vùng văn bản đã nhận dạng", expanded=True):
                for i, block in enumerate(text_blocks, 1):
                    st.markdown(f"**Vùng {i}:**")
                    st.code(block, language="text")
        
        # 2. Show merged text in document order (top-to-bottom, left-to-right)
        with st.expander("📝 Văn bản đã nhận dạng (đầy đủ)", expanded=False):
            # Get the text blocks with their positions from OCR results
            text_blocks = st.session_state.ocr_results.get(file_key, [])
            
            if not text_blocks:
                st.warning("Không có dữ liệu văn bản nào được tìm thấy.")
                st.stop(    )
            
            # Sort text blocks by their Y position (top-to-bottom) and then X position (left-to-right)
            sorted_blocks = sorted(
                [b for b in text_blocks if 'abs_box' in b and b.get('text', '').strip()],
                key=lambda b: (b['abs_box'][1], b['abs_box'][0])  # Sort by Y then X
            )
            
            # Group text into lines based on Y-coordinate (allowing some vertical tolerance)
            lines = []
            current_line = []
            last_y = None
            
            for block in sorted_blocks:
                y_pos = block['abs_box'][1]  # Top Y coordinate
                text = block['text'].strip()
                
                if not text:
                    continue
                    
                # If this block is on a new line (far enough below the last one)
                if last_y is not None and (y_pos - last_y) > 10:  # 10px tolerance
                    if current_line:
                        # Sort the current line by X coordinate before adding
                        current_line.sort(key=lambda b: b['abs_box'][0])
                        line_text = ' '.join([b['text'] for b in current_line])
                        lines.append(line_text)
                        current_line = []
                
                current_line.append(block)
                last_y = y_pos
            
            # Add the last line if it exists
            if current_line:
                current_line.sort(key=lambda b: b['abs_box'][0])
                line_text = ' '.join([b['text'] for b in current_line])
                lines.append(line_text)
            
            # Combine all lines with proper spacing
            formatted_text = '\n'.join(lines)
            
            # Display in a text area with improved formatting
            st.text_area(
                "Toàn bộ văn bản đã nhận dạng (sắp xếp theo thứ tự tài liệu):",
                value=formatted_text,
                height=400,
                key=f"full_text_{file_key}",
                help="Văn bản đã được sắp xếp theo thứ tự từ trên xuống dưới, từ trái sang phải"
            )
            
            # Save the full recognized text to a file in Output_Text_File directory
            try:
                # Ensure the output directory exists
                os.makedirs(Output_Text_File, exist_ok=True)
                
                # Create a filename based on the original file
                file_basename = os.path.splitext(os.path.basename(filename))[0]
                output_file = os.path.join(Output_Text_File, f"{file_basename}_full_text.txt")
                
                # Write the formatted text to the file
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(formatted_text)
                
            except Exception as e:
                st.error(f"Lỗi khi lưu file văn bản đầy đủ: {str(e)}")
        
        # 3. Show parsed data table if available
        display_df = pd.DataFrame()
        if hasattr(st.session_state, 'parsed_data') and file_key in st.session_state.parsed_data and st.session_state.parsed_data[file_key]:
            try:
                parsed_data = st.session_state.parsed_data[file_key]
                
                # Create DataFrame from parsed data
                if isinstance(parsed_data, dict):
                    display_df = pd.DataFrame([parsed_data])
                elif isinstance(parsed_data, list):
                    display_df = pd.DataFrame(parsed_data)
                
                if not display_df.empty:
                    with st.expander("📊 Thông tin đã trích xuất", expanded=True):
                        st.dataframe(display_df, use_container_width=True)
                        
                        # Add download button for the parsed data
                        excel_file = BytesIO()
                        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                            display_df.to_excel(writer, index=False, sheet_name='Thông tin hóa đơn')
                        
                        excel_file.seek(0)
                        st.download_button(
                            label="💾 Tải xuống file Excel",
                            data=excel_file,
                            file_name=f"ket_qua_phan_tich_{os.path.splitext(filename)[0]}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            except Exception as e:
                st.error(f"Lỗi khi hiển thị dữ liệu đã phân tích: {str(e)}")
        else:
            st.info("Không có dữ liệu đã phân tích để hiển thị")
        # --- Save Results ---
        output_dir = os.path.abspath(Output_Text_File)
        
        try:
            # Create output directory if it doesn't exist (with parents if needed)
            os.makedirs(output_dir, exist_ok=True)
            
            # Get the base name from the uploaded file (without extension)
            file_basename = os.path.splitext(os.path.basename(filename))[0]
            
            # Define the required columns in the specified order
            required_columns = [
                "Người mua",
                "MST người mua",
                "Người bán",
                "MST người bán",
                "Tên hàng hóa",
                "Đơn vị tính",
                "Số lượng",
                "Đơn giá",
                "Thành tiền",
                "Cộng tiền hàng",
                "Thuế GTGT",
                "Tổng cộng",
                "Đường dẫn ảnh"
            ]
            # Initialize with empty values for all required columns
            excel_data = {col: [""] * len(display_df) if not display_df.empty else [""] for col in required_columns}
            
            # Copy data from the display DataFrame to the new one
            if not display_df.empty:
                for col in display_df.columns:
                    if col in excel_data:
                        excel_data[col] = display_df[col].fillna("").values.tolist()
            
            # Create the final DataFrame with all required columns
            final_df = pd.DataFrame(excel_data)
            
            # Save Excel file with formatting
            excel_filename = os.path.join(output_dir, f"{file_basename}.xlsx")
            try:
                # Create a new workbook and select the active worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "OCR Results"
                
                # Define styles
                header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                cell_font = Font(name='Arial', size=11)
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                # Write headers with style
                for col_num, column in enumerate(required_columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Write data rows with style using final_df which has all required columns
                for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), 2):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.font = cell_font
                        cell.alignment = Alignment(horizontal='left')
                        
                        # Format numeric columns
                        if required_columns[c_idx-1] in ["Số lượng", "Đơn giá", "Thành tiền", "Cộng tiền hàng", "Thuế GTGT", "Tổng cộng"] and value != "":
                            try:
                                cell.number_format = '#,##0'
                            except:
                                pass
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
                # Save the workbook
                wb.save(excel_filename)
                st.success(f"Đã lưu file Excel: {os.path.basename(excel_filename)}")
                
            except Exception as e:
                st.error(f"Lỗi khi lưu file Excel: {str(e)}")
                # Fallback to a temporary directory if the main directory fails
                temp_dir = os.path.join(os.path.expanduser("~"), "temp_ocr_results")
                os.makedirs(temp_dir, exist_ok=True)
                excel_filename = os.path.join(temp_dir, f"{file_basename}.xlsx")
                display_df.to_excel(excel_filename, index=False, engine='openpyxl')
                st.warning(f"Đã lưu file Excel vào thư mục tạm: {excel_filename}")
            
            # Save TXT file with the full recognized text
            txt_filename = os.path.join(output_dir, f"{file_basename}.txt")
            try:
                # Get the formatted text from the full text section
                formatted_text = ""
                if hasattr(st.session_state, 'full_text') and file_key in st.session_state.full_text:
                    formatted_text = st.session_state.full_text[file_key]
                else:
                    # Fallback to original method if full text is not available
                    ocr_results = st.session_state.ocr_results.get(file_key, [])
                    blocks = []
                    for block in ocr_results:
                        if isinstance(block, dict) and 'text' in block:
                            text = block.get('text', '').strip()
                            if text:
                                blocks.append(text)
                        elif isinstance(block, str) and block.strip():
                            blocks.append(block.strip())
                    formatted_text = "\n".join(blocks)
                
                # Write the full text to the file
                with open(txt_filename, "w", encoding="utf-8") as f:
                    f.write(formatted_text)

                
                st.success(f"Đã lưu file văn bản: {os.path.basename(txt_filename)}")
            except Exception as e:
                st.error(f"Lỗi khi lưu file văn bản: {str(e)}")
                # Fallback to a temporary directory if the main directory fails
                temp_dir = os.path.join(os.path.expanduser("~"), "temp_ocr_results")
                os.makedirs(temp_dir, exist_ok=True)
                txt_filename = os.path.join(temp_dir, f"{file_basename}.txt")
                try:
                    with open(txt_filename, "w", encoding="utf-8") as f:
                        # Write each OCR block in order, separated by a divider
                        for i, block in enumerate(text_blocks, 1):
                            # Skip non-string blocks (like dictionaries)
                            if not isinstance(block, str):
                                continue
                                
                            block = str(block).strip()
                            if block:  # Only write non-empty blocks
                                f.write(f"--- Vùng {i} ---\n")
                                # Write each line of the block
                                for line in block.splitlines():
                                    line = line.strip()
                                    if line:  # Only write non-empty lines
                                        f.write(line + "\n")
                                f.write("\n")  # Add an extra newline between blocks
                except Exception as e2:
                    st.error(f"Lỗi khi lưu file văn bản dự phòng: {str(e2)}")
                
        except Exception as e:
            st.error(f"Lỗi khi lưu kết quả: {str(e)}")
            # Try to use a temporary directory as a last resort
            temp_dir = os.path.join(os.path.expanduser("~"), "temp_ocr_results")
            os.makedirs(temp_dir, exist_ok=True)
            
            # Save Excel to temp directory
            excel_filename = os.path.join(temp_dir, f"{file_basename}.xlsx")
            if 'display_df' in locals() and not display_df.empty:
                display_df.to_excel(excel_filename, index=False, engine='openpyxl')
            else:
                # Create a simple DataFrame with the full text if no parsed data is available
                pd.DataFrame({'Văn bản đã nhận dạng': [full_text]}).to_excel(excel_filename, index=False)
            
            # Save TXT to temp directory with top-to-bottom order
            txt_filename = os.path.join(temp_dir, f"{file_basename}.txt")
            try:
                # Get the original OCR results with coordinates
                ocr_results = st.session_state.ocr_results.get(file_key, [])
                
                # Extract text blocks with their Y-coordinates
                blocks_with_pos = []
                for i, block in enumerate(ocr_results):
                    if isinstance(block, dict) and 'box' in block and 'text' in block:
                        # Get the top Y-coordinate of the bounding box
                        y_coord = min(point[1] for point in block['box'])
                        blocks_with_pos.append({
                            'y': y_coord,
                            'text': block['text']
                        })
                    elif isinstance(block, str):
                        # For string blocks, use their position in the list as a fallback
                        blocks_with_pos.append({
                            'y': i * 100,  # Space them out by 100 units
                            'text': block
                        })
                
                # Sort blocks by Y-coordinate (top to bottom)
                sorted_blocks = sorted(blocks_with_pos, key=lambda x: x['y'])
                
                # Write the sorted blocks to the file
                with open(txt_filename, "w", encoding="utf-8") as f:
                    for i, block_info in enumerate(sorted_blocks, 1):
                        block = str(block_info['text']).strip()
                        if block:  # Only write non-empty blocks
                            # Write the block header
                            f.write(f"--- Vùng {i} ---\n")
                            # Write each line of the block
                            for line in block.splitlines():
                                line = line.strip()
                                if line:  # Only write non-empty lines
                                    f.write(line + "\n")
                            f.write("\n")  # Add an extra newline between blocks
            except Exception as e:
                st.error(f"Lỗi khi lưu file văn bản tạm thời: {str(e)}")

            # Show the full recognized text in an expander with clean formatting
            with st.expander("📄 Xem toàn bộ văn bản đã nhận dạng", expanded=False):
                # Get the parsed data for this file
                parsed_data = st.session_state.parsed_data.get(file_key, {})
                
                # Create a structured output
                output = []
                
                # 1. Header Section
                output.append("=" * 80)
                output.append("HÓA ĐƠN GIÁ TRỊ GIA TĂNG".center(80))
                output.append("=" * 80)
                output.append("")
                
                # 2. Process text blocks grouped by YOLO regions
                current_region = -1
                region_count = 0
                
                # Get the text blocks from OCR results
                text_blocks = st.session_state.ocr_results.get(file_key, [])
                
                for i, block in enumerate(text_blocks):
                    # Check if this is a region separator
                    if block.get('is_region_separator', False):
                        output.append("-" * 80)
                        continue
                        
                    # Get YOLO region info
                    region_idx = block.get('yolo_region', -1)
                    
                    # If this is a new region, add a header
                    if region_idx != current_region:
                        current_region = region_idx
                        region_count += 1
                        output.append(f"\nVÙNG VĂN BẢN {region_count}:")
                        output.append("-" * 40)
                
                # Add a summary of detected regions
                if region_count > 0:
                    output.append("\n" + "=" * 80)
                    output.append(f"Đã phát hiện {region_count} vùng văn bản trong hình ảnh.")
                
                # 3. Add structured data if available
                if parsed_data:
                    output.append("\nTHÔNG TIN ĐÃ ĐƯỢC XỬ LÝ:")
                    output.append("=" * 80)
                    
                    # Add seller information
                    output.append("\nBÊN BÁN:")
                    output.append(f"- Tên công ty: {parsed_data.get('ten_cong_ty', 'N/A')}")
                    output.append(f"- Mã số thuế: {parsed_data.get('ma_so_thue', 'N/A')}")
                    output.append(f"- Địa chỉ: {parsed_data.get('dia_chi', 'N/A')}")
                    
                    # Add invoice details
                    output.append("\nTHÔNG TIN HÓA ĐƠN:")
                    output.append(f"- Số hóa đơn: {parsed_data.get('so_hoa_don', 'N/A')}")
                    output.append(f"- Ngày lập: {parsed_data.get('ngay_lap', 'N/A')}")
                    output.append(f"- Tổng thanh toán: {parsed_data.get('tong_cong', 'N/A')} VNĐ")
                
                # Join all lines with newlines
                formatted_text = '\n'.join(output)
                
                # Display in a text area with monospace font for better alignment
                st.text_area(
                    "Văn bản đã nhận dạng (nhóm theo vùng):",
                    formatted_text,
                    height=500,
                    key=f"formatted_text_{file_key}",
                    help="Văn bản đã được nhóm theo các vùng phát hiện bởi YOLO"
                )
                
                # Show the original text blocks in a collapsed section
                with st.expander("Xem chi tiết các khối văn bản", expanded=False):
                    st.json({
                        'text_blocks': text_blocks,
                        'full_text': full_text
                    }, expanded=False)
            
            if file_key in st.session_state.ocr_results:
                del st.session_state.ocr_results[file_key]
            if file_key in st.session_state.df_results:
                del st.session_state.df_results[file_key]
            if file_key in st.session_state.full_texts:
                del st.session_state.full_texts[file_key]
            st.rerun()